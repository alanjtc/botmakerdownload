import csv
import re
from datetime import datetime
from tkinter import Tk, filedialog
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from datetime import datetime
from tkinter import Tk, filedialog
VALORES_AUTENTICACION = ["Autenticación Exitosa", "Contacto No Vinculado A Proveedor", "No Concluida"]
VALORES_RESULTADO = ["Entrega Exitosa", "No encontró información"]
SERVICIOS_VALIDOS = [ 'Actualiza Ariba','SBN Órdenes','SBN Facturas','Cert. Tributarios','Facturas por pagar','Aviso de pagos','Relación Comercial','Manual de procesos','Eventos Mercantiles']
SERVICIOS_CON_SOCIEDAD = ['Cert. Tributarios','Facturas por pagar','Aviso de pagos','Relación Comercial',]

SOCIEDADES_VALIDAS = [
    'GeoPark',
    'Amerisur'
]

def procesar_conversaciones(archivo_txt, archivo_xlsx):
    # El archivo original tenía 8 columnas separadas por ";".
    # A veces ahora se inserta una columna adicional para la plataforma
    # (ej. "WhatsApp") entre el ID del contacto y la fecha.  El regex
    # a continuación usa un grupo sin captura opcional para descartar ese
    # campo si está presente, de modo que `creationTime` siempre sea el
    # valor correcto.
    patron = re.compile(r"^(.*?);(.*?);(?:.*?;)?(.*?);(.*?);(.*?);(.*?);(.*?);(.*)$")
    conversaciones = {}
    links_chat = {}

    with open(archivo_txt, encoding="utf-8") as f:
        for linea in f:
            match = patron.match(linea.strip())
            if match:
                chatId, platformContactId, creationTime, sender, msg, extra, extra2, url = match.groups()
                fecha_dia = creationTime[:10]
                clave = f"{chatId}_{fecha_dia}"
                if clave not in conversaciones:
                    conversaciones[clave] = []
                    links_chat[clave] = url if url.startswith("http") else ""
                
                conversaciones[clave].append({
                    "chatId": chatId,
                    "platformContactId": platformContactId,
                    "creationTime": creationTime,
                    "sender": sender,
                    "msg": msg.strip(),
                    "extra": extra.strip(),
                    "extra2": extra2.strip(),
                    "url": url
                })

    filas = []
    autenticacion_por_numero = {}  # Diccionario para autenticación por número
    nit_por_numero = {}  # Diccionario para NIT

    # Recorre las conversaciones agrupadas por clave
    for clave, mensajes in conversaciones.items():
        platform_id = mensajes[0]["platformContactId"]
        autenticacion = autenticacion_por_numero.get(platform_id, "No")
        nit_actual = nit_por_numero.get(platform_id, "")  # Usa el último NIT conocido o vacío
        bloques = []
        bloque_actual = None
        chatId = mensajes[0]["chatId"]
        ultima_pregunta_servicio = False

        # Si es conversación de prueba, omitirla (no agregar fila)
        if "geoparkbot_test_chat" in mensajes[0]["platformContactId"]:
            continue  # Omitir conversaciones de prueba

        # Procesa los mensajes de la conversación
        i = 0
        sociedad_temp = ""
        
        while i < len(mensajes):
            m = mensajes[i]
            texto_total = f"{m['msg']} {m['extra']} {m['extra2']}".strip().lower()

            if "te damos la bienvenida" in texto_total:
                autenticacion = "Sí"

            # Captura sociedad solo si el servicio está en SERVICIOS_CON_SOCIEDAD
            if m["sender"] == "user" and any(s in texto_total for s in ["geopark", "amerisur"]):
                sociedad_temp = m["msg"] or m["extra2"] or m["extra"]
                if bloque_actual and bloque_actual["servicio"] in SERVICIOS_CON_SOCIEDAD:
                    bloque_actual["sociedad"] = sociedad_temp

            if m["sender"] == "user" and re.fullmatch(r"\d{5,}", m["msg"].strip()):
                nit_valor = m["msg"].strip()
                nit_actual = nit_valor  # Actualiza el NIT actual para este número
                if bloque_actual:
                    bloque_actual["nit"] = nit_valor
                else:
                    bloques.append({
                        "servicio": "",
                        "sociedad": sociedad_temp if bloque_actual and bloque_actual["servicio"] in SERVICIOS_CON_SOCIEDAD else "",
                        "resultado": "No Encontrado",
                        "documento": "",
                        "correos": "",
                        "nit": nit_valor
                    })

            if m["sender"] == "bot" and "¿en cuál de las siguientes tareas te puedo ayudar?" in texto_total:
                ultima_pregunta_servicio = True
                i += 1
                continue

            if ultima_pregunta_servicio and m["sender"] == "user":
                servicio_valor = m["msg"] or m["extra2"] or m["extra"]
                # Validar si el servicio está en la lista de servicios válidos
                if any(servicio in servicio_valor for servicio in SERVICIOS_VALIDOS):
                    servicio_final = servicio_valor
                else:
                    servicio_final = ""  # O puedes poner "Servicio no válido" si prefieres

                bloque_actual = {
                    "servicio": servicio_final,
                    "sociedad": sociedad_temp if servicio_final in SERVICIOS_CON_SOCIEDAD else "",
                    "resultado": "No Encontrado",
                    "documento": "",
                    "correos": "",
                    "nit": ""
                }
                bloques.append(bloque_actual)
                ultima_pregunta_servicio = False
                i += 1
                continue

            # Si la sociedad se menciona después de escoger servicio, también la asigna
            if bloque_actual and m["sender"] == "user" and any(s in texto_total for s in ["geopark", "amerisur"]):
                if bloque_actual["servicio"] in SERVICIOS_CON_SOCIEDAD:
                    bloque_actual["sociedad"] = m["msg"] or m["extra2"] or m["extra"]

            # Buscar documento PDF y extraer correos de la siguiente línea si corresponde
            documento_encontrado = False
            for campo in [m["msg"], m["extra"], m["extra2"], m["url"]]:
                if bloque_actual and ".pdf" in campo.lower():
                    bloque_actual["documento"] = campo
                    bloque_actual["resultado"] = "Encontrado"
                    documento_encontrado = True

            if documento_encontrado and i + 1 < len(mensajes):
                siguiente = mensajes[i + 1]
                texto_sig = f"{siguiente['msg']} {siguiente['extra']} {siguiente['extra2']}".strip()
                if "El PDF fue enviado a los siguientes correos" in texto_sig:
                    texto_correos = texto_sig.replace("-", " ")
                    correos_encontrados = re.findall(r"[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+", texto_correos)
                    if correos_encontrados:
                        bloque_actual["correos"] = "; ".join(sorted(set(correos_encontrados)))

            # Si no se llenó correos, buscar en los campos actuales
            if bloque_actual and not bloque_actual["correos"]:
                correos_encontrados = []
                for campo in [m["msg"], m["extra"], m["extra2"]]:
                    if campo:
                        correos_encontrados += re.findall(r"[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+", campo)
                if correos_encontrados:
                    bloque_actual["correos"] = "; ".join(sorted(set(correos_encontrados)))
            i += 1

        if not bloques:
            bloques.append({
                "servicio": "Solo Saludo",
                "sociedad": "",
                "resultado": "No Encontrado",
                "documento": "",
                "correos": "",
                "nit": nit_actual  # Usa el último NIT conocido
            })

        # Si algún bloque no tiene NIT, usa el último conocido
        for b in bloques:
            if not b.get("nit"):
                b["nit"] = nit_actual

        # Intentamos convertir la fecha de creación; si no es un ISO válido
        # (p.ej. quedó la plataforma en vez del timestamp), omitimos la
        # conversación completa y avisamos en consola.
        try:
            dt = datetime.fromisoformat(mensajes[0]["creationTime"].replace("Z", "+00:00"))
        except Exception as e:
            print(f"⚠️ Fecha inválida para clave {clave}: {mensajes[0]['creationTime']} (se omite conversación)")
            continue
        fecha = dt.strftime("%Y-%m-%d")
        hora = dt.strftime("%H:%M:%S")

        for b in bloques:
            # Determina el valor base de autenticación ("Sí"/"No"/"")
            if b["documento"]:
                base_autenticacion = "Sí"
            elif b["servicio"] in SERVICIOS_CON_SOCIEDAD:
                base_autenticacion = "Sí"
            else:
                base_autenticacion = autenticacion  # valor previo por número ("Sí", "No", etc.)

            # Si es "Solo Saludo" la columna autenticación queda vacía
            if b["servicio"] == "Solo Saludo":
                valor_autenticacion = VALORES_AUTENTICACION[2]
            else:
                # Normalizar y mapear valores (evita valores vacíos)
                ba = (base_autenticacion or "").strip().lower()
                if ba in ("sí", "si", "s", "yes", "y"):
                    valor_autenticacion = VALORES_AUTENTICACION[0]  # "Autenticación Exitosa"
                elif ba in ("no", "n", "false", "0"):
                    valor_autenticacion = VALORES_AUTENTICACION[1]  # "Contacto No Vinculado A Proveedor"
                else:
                    # Valor por defecto cuando no hay información clara
                    valor_autenticacion = VALORES_AUTENTICACION[2]  # "No Concluida"

            # Si el resultado es "No Encontrado", no llenar correos enviados
            if b["resultado"] == "No Encontrado":
                b["correos"] = ""

            # Mapear valor de resultado a VALORES_RESULTADO
            if b.get("resultado") == "Encontrado":
                resultado_mostrar = VALORES_RESULTADO[0]  # "Entrega Exitosa"
            else:
                resultado_mostrar = VALORES_RESULTADO[1]  # "No encontró información"

            # Validar que la sociedad esté en el listado de sociedades válidas
            sociedad_valida = b["sociedad"] if b["sociedad"] in SOCIEDADES_VALIDAS else ""

            # Si autenticación es "Autenticación Exitosa" pero no hay servicio, no mostrar la conversación
            if valor_autenticacion == VALORES_AUTENTICACION[0] and not (b.get("servicio") and b.get("servicio").strip()):
                continue

            # Solo se modifica el valor mostrado en la columna "Servicio" si está vacío;
            # no se modifica el valor de autenticación.
            if not (b.get("servicio") and b.get("servicio").strip()):
                servicio_mostrar = VALORES_AUTENTICACION[1]
            else:
                servicio_mostrar = b["servicio"]

            filas.append([
                fecha,
                hora,
                chatId,
                mensajes[0]["platformContactId"],
                b.get("nit", ""),
                valor_autenticacion,
                servicio_mostrar,
                sociedad_valida,
                resultado_mostrar,
                b["documento"],
                b["correos"],
                links_chat.get(clave, "")
            ])

        # Si en algún bloque se generó un PDF, la autenticación será "Sí"
        if any(".pdf" in (b.get("documento") or "").lower() for b in bloques):
            autenticacion = "Sí"

        # Guarda el valor de autenticación y NIT para este número para futuras conversaciones
        autenticacion_por_numero[platform_id] = autenticacion
        nit_por_numero[platform_id] = nit_actual

    # Crear archivo XLSX directamente
    wb = openpyxl.Workbook()
    ws = wb.active

    # Escribir encabezados
    encabezados = [
        "Fecha", "Hora", "ID Conversación", "Celular Usuario", "NIT Proveedor",
        "¿Autenticación Exitosa?", "Servicio Consultado", "Sociedad", "Resultado",
        "Documento Generado", "Correos Enviados", "Link Conversación"
    ]
    ws.append(encabezados)

    # Escribir filas
    for fila in filas:
        ws.append(fila)

    # Formatear encabezados
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Ajustar ancho de columnas automáticamente
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Alinear el resto de las celdas
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(archivo_xlsx)
    print(f"✅ XLSX generado exitosamente en: {archivo_xlsx}")


if __name__ == "__main__":
    print("🚀 Procesador de conversaciones de bot a XLSX")
    root = Tk()
    root.withdraw()
    archivo_txt = filedialog.askopenfilename(title="Selecciona el archivo .txt de entrada", filetypes=[("Archivos de texto", "*.txt")])
    if not archivo_txt:
        print("❌ No se seleccionó archivo de entrada.")
        root.destroy()
        exit()

    archivo_xlsx = filedialog.asksaveasfilename(title="Guardar archivo XLSX", defaultextension=".xlsx", filetypes=[("Archivo XLSX", "*.xlsx")])
    if not archivo_xlsx:
        print("❌ No se seleccionó ubicación de salida.")
        root.destroy()
        exit()

    procesar_conversaciones(archivo_txt, archivo_xlsx)
    root.destroy()