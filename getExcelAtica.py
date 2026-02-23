import csv
import re
from datetime import datetime
from tkinter import Tk, filedialog
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

VALORES_AUTENTICACION = ["Autenticación Exitosa", "Contacto No Vinculado A Proveedor", "No Concluida"]
VALORES_RESULTADO = ["Entrega Exitosa", "No encontró información"]

# Servicios válidos para Atica
SERVICIOS_VALIDOS = ['Certificados tributarios', 'Estado de cuenta', 'Información General', 'Estatus de facturas', 'Detalle de últimos pagos']

# Servicios que requieren selección de sociedad
SERVICIOS_CON_SOCIEDAD = ['Certificados tributarios', 'Estado de cuenta', 'Estatus de facturas', 'Detalle de últimos pagos', "Información General"]

# Sociedades del Grupo Atica
SOCIEDADES_VALIDAS = [
    'IA4 SAS',
    'VALREX SAS',
    'INDUSTRIA AMBIENTAL SAS',
    'ECOENTORNO',
    'LUBRYESP SAS',
    'BIOLOGICOS Y CONTAMINADOS',
    'ECOLOGISTICA',
    'ASESORIAS Y SERVICIOS ECO'
]

# Mapeo de números a nombres de sociedad
MAPEO_SOCIEDAD = {
    '1': 'IA4 SAS',
    '2': 'VALREX SAS',
    '3': 'INDUSTRIA AMBIENTAL SAS',
    '4': 'ECOENTORNO',
    '5': 'LUBRYESP SAS',
    '6': 'BIOLOGICOS Y CONTAMINADOS',
    '7': 'ECOLOGISTICA',
    '8': 'ASESORIAS Y SERVICIOS ECO'
}

def extraer_sociedades_de_mensaje(mensaje):
    """
    Extrae las sociedades del mensaje del bot que lista las opciones.
    Retorna un diccionario {número: nombre_sociedad}
    """
    mapeo = {}
    # Buscar líneas con formato "N. NOMBRE" donde N es un número
    # Usar split de espacios múltiples para separar opciones
    patron_sociedad = re.compile(r'(\d+)\.\s+([A-Z\s&0-9]+?)(?=\s{2,}|\s\d+\.|$)')
    
    matches = patron_sociedad.findall(mensaje)
    for num, nombre in matches:
        mapeo[num] = nombre.strip()
    
    return mapeo


def procesar_conversaciones(archivo_txt, archivo_xlsx):
    # Patrón para parsear las líneas del archivo
    patron = re.compile(r"^(.*?);(.*?);(.*?);(.*?);(.*?);(.*?);(.*?);(.*)$")
    conversaciones = {}

    with open(archivo_txt, encoding="utf-8") as f:
        for linea in f:
            match = patron.match(linea.strip())
            if match:
                chatId, platformContactId, plataforma, creationTime, sender, msg, extra, extra2 = match.groups()
                fecha_dia = creationTime[:10]
                clave = f"{chatId}_{fecha_dia}"
                
                if clave not in conversaciones:
                    conversaciones[clave] = []
                
                conversaciones[clave].append({
                    "chatId": chatId,
                    "platformContactId": platformContactId,
                    "plataforma": plataforma,
                    "creationTime": creationTime,
                    "sender": sender,
                    "msg": msg.strip().rstrip(';'),
                    "extra": extra.strip().rstrip(';'),
                    "extra2": extra2.strip().rstrip(';'),
                })

    filas = []
    autenticacion_por_numero = {}
    nit_por_numero = {}

    for clave, mensajes in conversaciones.items():
        platform_id = mensajes[0]["platformContactId"]
        autenticacion = autenticacion_por_numero.get(platform_id, "No")
        nit_actual = nit_por_numero.get(platform_id, "")
        bloques = []
        bloque_actual = None
        chatId = mensajes[0]["chatId"]
        ultima_pregunta_servicio = False
        sociedad_temp = ""
        mapeo_sociedades_dinamico = {}  # Para almacenar el mapeo dinámico del chat

        # Omitir conversaciones de prueba
        if "test" in mensajes[0]["platformContactId"].lower():
            continue

        i = 0
        while i < len(mensajes):
            m = mensajes[i]
            texto_total = f"{m['msg']} {m['extra']} {m['extra2']}".strip().lower()

            # Capturar mapeo dinámico de sociedades cuando el bot lista las opciones
            if m["sender"] == "bot" and "selecciona la sociedad" in texto_total:
                msg_completo = f"{m['msg']} {m['extra']} {m['extra2']}"
                mapeo_sociedades_dinamico = extraer_sociedades_de_mensaje(msg_completo)

            # Detectar autenticación exitosa
            if "sí" in texto_total or "autorización" in texto_total:
                if m["sender"] == "user":
                    if "sí" in m['msg'].lower() or "sí" in m['extra2'].lower():
                        autenticacion = "Sí"

            # Capturar NIT
            if m["sender"] == "user" and re.fullmatch(r"\d{5,}", m["msg"].strip()):
                nit_valor = m["msg"].strip()
                nit_actual = nit_valor
                if bloque_actual:
                    bloque_actual["nit"] = nit_valor
                else:
                    bloques.append({
                        "servicio": "",
                        "sociedad": "",
                        "resultado": "No Encontrado",
                        "documento": "",
                        "correos": "",
                        "nit": nit_valor
                    })

            # Detectar pregunta de servicio
            if m["sender"] == "bot" and "¿qué necesitas" in texto_total:
                ultima_pregunta_servicio = True
                i += 1
                continue

            # Procesar respuesta de servicio del usuario (cuando hay pregunta explícita)
            if ultima_pregunta_servicio and m["sender"] == "user":
                servicio_valor = m["msg"] or m["extra2"] or m["extra"]
                # Validar si el servicio está en la lista válida
                if any(servicio.lower() in servicio_valor.lower() for servicio in SERVICIOS_VALIDOS):
                    servicio_final = servicio_valor
                else:
                    servicio_final = ""

                bloque_actual = {
                    "servicio": servicio_final,
                    "sociedad": "",
                    "resultado": "No Encontrado",
                    "documento": "",
                    "correos": "",
                    "nit": ""
                }
                bloques.append(bloque_actual)
                ultima_pregunta_servicio = False
                i += 1
                continue

            # Detectar cuando el usuario menciona directamente un servicio (sin pregunta explícita)
            if m["sender"] == "user" and not ultima_pregunta_servicio:
                texto_msg = f"{m['msg']} {m['extra2']}".lower().strip()
                servicio_encontrado = None
                for servicio in SERVICIOS_VALIDOS:
                    if servicio.lower() in texto_msg:
                        servicio_encontrado = servicio
                        break
                
                # Si se encontró un servicio válido
                if servicio_encontrado:
                    # Si hay bloque_actual sin servicio, actualizar ese bloque
                    if bloque_actual and not bloque_actual.get("servicio"):
                        bloque_actual["servicio"] = servicio_encontrado
                    # Si no hay bloque o el actual ya tiene servicio, crear uno nuevo
                    elif not bloque_actual or bloque_actual.get("servicio"):
                        bloque_actual = {
                            "servicio": servicio_encontrado,
                            "sociedad": "",
                            "resultado": "No Encontrado",
                            "documento": "",
                            "correos": "",
                            "nit": ""
                        }
                        bloques.append(bloque_actual)

            # Capturar sociedad cuando se menciona
            if m["sender"] == "user" and bloque_actual:
                usuario_input = m["msg"].strip()
                # Primero intentar mapeo dinámico (extraído del chat)
                if usuario_input in mapeo_sociedades_dinamico:
                    sociedad_seleccionada = mapeo_sociedades_dinamico[usuario_input]
                    if bloque_actual["servicio"] in SERVICIOS_CON_SOCIEDAD:
                        bloque_actual["sociedad"] = sociedad_seleccionada
                # Luego intentar mapeo numérico fijo (1-8)
                elif usuario_input in MAPEO_SOCIEDAD:
                    sociedad_seleccionada = MAPEO_SOCIEDAD[usuario_input]
                    if bloque_actual["servicio"] in SERVICIOS_CON_SOCIEDAD:
                        bloque_actual["sociedad"] = sociedad_seleccionada
                # Luego buscar por nombre de sociedad
                elif not bloque_actual["sociedad"]:  # Solo si aún no tiene sociedad
                    for sociedad in SOCIEDADES_VALIDAS:
                        if sociedad.lower() in texto_total or sociedad.split()[0].lower() in texto_total:
                            if bloque_actual["servicio"] in SERVICIOS_CON_SOCIEDAD:
                                bloque_actual["sociedad"] = sociedad
                                break

            # Buscar documento PDF
            documento_encontrado = False
            for campo in [m["msg"], m["extra"], m["extra2"]]:
                if bloque_actual and ".pdf" in campo.lower():
                    bloque_actual["documento"] = campo
                    bloque_actual["resultado"] = "Encontrado"
                    documento_encontrado = True

            # Buscar correos
            if bloque_actual and not bloque_actual["correos"]:
                correos_encontrados = []
                for campo in [m["msg"], m["extra"], m["extra2"]]:
                    if campo:
                        correos_encontrados += re.findall(r"[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+", campo)
                
                if correos_encontrados:
                    bloque_actual["correos"] = "; ".join(sorted(set(correos_encontrados)))
                elif "correo registrado" in texto_total or "pagos@atica.co" in texto_total:
                    bloque_actual["correos"] = "pagos@atica.co"

            i += 1

        # Si no hay bloques, crear uno por defecto
        if not bloques:
            bloques.append({
                "servicio": "Solo Saludo",
                "sociedad": "",
                "resultado": "No Encontrado",
                "documento": "",
                "correos": "",
                "nit": nit_actual
            })

        # Asignar NIT a bloques sin NIT
        for b in bloques:
            if not b.get("nit"):
                b["nit"] = nit_actual

        # Procesar fecha y hora
        dt = datetime.fromisoformat(mensajes[0]["creationTime"].replace("Z", "+00:00"))
        fecha = dt.strftime("%Y-%m-%d")
        hora = dt.strftime("%H:%M:%S")

        # Generar filas del Excel
        for b in bloques:
            # Determinar valor de autenticación
            if b["documento"]:
                base_autenticacion = "Sí"
            elif b["servicio"] in SERVICIOS_CON_SOCIEDAD:
                base_autenticacion = "Sí"
            else:
                base_autenticacion = autenticacion

            if b["servicio"] == "Solo Saludo":
                valor_autenticacion = VALORES_AUTENTICACION[2]
            else:
                ba = (base_autenticacion or "").strip().lower()
                if ba in ("sí", "si", "s", "yes", "y"):
                    valor_autenticacion = VALORES_AUTENTICACION[0]
                elif ba in ("no", "n", "false", "0"):
                    valor_autenticacion = VALORES_AUTENTICACION[1]
                else:
                    valor_autenticacion = VALORES_AUTENTICACION[2]

            # Si resultado es "No Encontrado", limpiar correos
            if b["resultado"] == "No Encontrado":
                b["correos"] = ""

            # Mapear resultado
            if b.get("resultado") == "Encontrado":
                resultado_mostrar = VALORES_RESULTADO[0]
            else:
                resultado_mostrar = VALORES_RESULTADO[1]

            # Validar sociedad
            sociedad_valida = b["sociedad"] if b["sociedad"] else ""

            # Si autenticación exitosa pero sin servicio, y sin documento, omitir
            if valor_autenticacion == VALORES_AUTENTICACION[0] and not (b.get("servicio") and b.get("servicio").strip()) and not b.get("documento"):
                continue

            # Mostrar servicio o "Contacto No Vinculado"
            if not (b.get("servicio") and b.get("servicio").strip()):
                servicio_mostrar = VALORES_AUTENTICACION[1]
            else:
                servicio_mostrar = b["servicio"]

            filas.append([
                fecha,
                hora,
                chatId,
                mensajes[0]["platformContactId"],
                b.get("nit", ""),
                valor_autenticacion,
                servicio_mostrar,
                sociedad_valida,
                resultado_mostrar,
                b["documento"],
                b["correos"],
                mensajes[0]["plataforma"]
            ])

        # Actualizar autenticación para futuras conversaciones
        if any(".pdf" in (b.get("documento") or "").lower() for b in bloques):
            autenticacion = "Sí"

        autenticacion_por_numero[platform_id] = autenticacion
        nit_por_numero[platform_id] = nit_actual

    # Crear archivo XLSX
    wb = openpyxl.Workbook()
    ws = wb.active

    # Escribir encabezados
    encabezados = [
        "Fecha", "Hora", "ID Conversación", "Celular Usuario", "NIT Proveedor",
        "¿Autenticación Exitosa?", "Servicio Consultado", "Sociedad", "Resultado",
        "Documento Generado", "Correos Enviados", "Plataforma"
    ]
    ws.append(encabezados)

    # Escribir filas
    for fila in filas:
        ws.append(fila)

    # Formatear encabezados
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Alinear celdas
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(archivo_xlsx)
    print(f"✅ XLSX generado exitosamente en: {archivo_xlsx}")


if __name__ == "__main__":
    print("🚀 Procesador de conversaciones Atica a XLSX")
    root = Tk()
    root.withdraw()
    
    archivo_txt = filedialog.askopenfilename(
        title="Selecciona el archivo .txt de entrada", 
        filetypes=[("Archivos de texto", "*.txt")]
    )
    if not archivo_txt:
        print("❌ No se seleccionó archivo de entrada.")
        root.destroy()
        exit()

    archivo_xlsx = filedialog.asksaveasfilename(
        title="Guardar archivo XLSX", 
        defaultextension=".xlsx", 
        filetypes=[("Archivo XLSX", "*.xlsx")]
    )
    if not archivo_xlsx:
        print("❌ No se seleccionó ubicación de salida.")
        root.destroy()
        exit()

    procesar_conversaciones(archivo_txt, archivo_xlsx)
    root.destroy()